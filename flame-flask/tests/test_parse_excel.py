import pandas as pd

# 确保已安装 pandas 和 openpyxl
# 读取 Excel 文件
try:
    df = pd.read_excel('/Users/xcx/WorkSpaces/BlueFlame/flame-flask/static/func_case_template.xlsx', engine='openpyxl')
    # 将每行数据转换为对象
    data = []
    for index, row in df.iterrows():
        obj = {}
        for column in df.columns:
            obj[column] = row[column]
        data.append(obj)
    # 查看数据
    for obj in data:
        print(obj)
        print('-' * 80)
except Exception as e:
    print("Error occurred while reading Excel file:", str(e))

# 根据需要处理数据...
# （可选）将数据写回到 Excel 文件
# df.to_excel('path_to_your_new_file.xlsx', engine='openpyxl')

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any

def parse_case_template_excel(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]: # type: ignore
    # 加载Excel工作簿
    workbook = openpyxl.load_workbook(file_path)
    
    # 选择工作表
    if sheet_name:
        sheet: Worksheet = workbook[sheet_name]
    else:
        sheet: Worksheet = workbook.active  # type: ignore
    

    # 获取表头
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
    # 定义中英文对应的字典

    header_dict = {
        '用例编号': 'case_id',
        '用例名称': 'case_name',
        '测试环境': 'test_environment',
        '所属模块': 'module',
        '前置条件': 'precondition',
        '操作步骤': 'steps',
        '预期结果': 'expected_result',
        '测试结果': 'test_result',
        '创建时间': 'create_time',
        '修改时间': 'modify_time'
    }
    data = []
    # 从第二行开始读取数据
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row_data = {}
        for cell, header in zip(row, headers):
            cell_value = cell.value
            for merged_cell_range in sheet.merged_cells: # type: ignore
                if cell.coordinate in merged_cell_range:
                    cell_value = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value
                    break
            # 将中文表头转换为英文表头
            english_header = header_dict.get(header, header)
            row_data[english_header] = cell_value
        data.append(row_data)
    
    return data
file_path = '/Users/xcx/WorkSpaces/BlueFlame/flame-flask/static/func_case_template.xlsx'
data = read_merged_excel(file_path)

# 显示前几行数据
for item in data[:5]:
    print(item)
    
