import io
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.datastructures import FileStorage
from flask import current_app


class CaseTemplate:

    def __init__(self,
                 file: FileStorage,
                 user_id: str,
                 project_id: int,
                 case_type: str = ""):
        self.file = file
        self.user_id = user_id
        self.project_id = project_id
        self.case_type = case_type
        self.data = self.parse_case_template_excel()

    def __repr__(self) -> str:
        return f"user_id: {self.user_id}, project_id: {self.project_id}, case_type: {self.case_type}, data: {self.data}"

    # def save_case_template(self):
    #     # 保存用例模板到数据库
    #     # ...
    #     current_app.logger.info(f"保存用例模板成功")

    # def update_case_template(self):
    #     # 更新用例模板到数据库
    #     # ...
    #     current_app.logger.info(f"更新用例模板成功")

    # def delete_case_template(self):
    #     # 删除用例模板
    #     # ...
    #     current_app.logger.info(f"删除用例模板成功")

    # def get_case_template(self):
    #     # 获取用例模板
    #     # ...
    #     current_app.logger.info(f"获取用例模板成功")

    def to_dict(self):
        return {
            "file": self.file,
            "user_id": self.user_id,
            "project_id": self.project_id,
            "case_type": self.case_type,
            "data": self.data
        }

    def get_data(self):
        return self.data

    def parse_case_template_excel(self,
                                  sheet_name: str = ""
                                  ) -> List[Dict[str, Any]]:
        try:
            file_stream = io.BytesIO(self.file.read())
            workbook = load_workbook(file_stream, data_only=True)
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            headers = [
                cell.value
                for cell in next(sheet.iter_rows(min_row=1, max_row=1)) # type: ignore
            ]
            header_dict = self.get_header_dict()
            data = self.extract_data(sheet, headers, header_dict) # type: ignore
            return data
        except Exception as e:
            print(f"解析模板时发生错误: {str(e)}")
            return []

    def get_header_dict(self) -> Dict[str, str]:
        return {
            '用例编号': 'case_id_by_user',
            '用例名称': 'case_name',
            '测试环境': 'test_environment',
            '所属模块': 'module',
            '前置条件': 'precondition',
            '操作步骤': 'steps',
            '预期结果': 'expected_result',
            '测试结果': 'test_result',
            '创建时间': 'create_time',
            '修改时间': 'modify_time'
        }

    def extract_data(self, sheet: Worksheet, headers: List[str],
                     header_dict: Dict[str, str]) -> List[Dict[str, Any]]:
        data = []
        # 获取合并单元格的值
        merged_cells_value = self.get_merged_cells_value(sheet)
        # 遍历工作表中的所有行
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            row_data = {}
            # 遍历当前行的所有单元格
            for cell, header in zip(row, headers):
                # 获取单元格的值，如果存在合并单元格，则获取合并单元格的值
                cell_value = merged_cells_value.get(cell.coordinate,
                                                    cell.value)
                english_header = header_dict.get(header, header)
                # 如果英文表头不在row_data中，则添加到row_data中
                if english_header not in row_data:
                    row_data[
                        english_header] = cell_value if cell_value is not None else ""
                else:
                    row_data[
                        english_header] = cell_value if cell_value is not None else row_data[
                            english_header]
                # 判断必填字段, 筛选出错误用例, 添加标识
                if english_header in ["module", "case_name", "precondition"
                                      ] and row_data[english_header] == "":
                    row_data["dirty"] = True
                    row_data[english_header] = "required"
            data.append(row_data)
        return data

    def get_merged_cells_value(self, sheet: Worksheet) -> Dict[str, Any]:
        merged_cells_value = {}
        for merged_range in sheet.merged_cells.ranges: # type: ignore
            top_left_value = sheet.cell(merged_range.min_row,
                                        merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col,
                                 merged_range.max_col + 1):
                    merged_cells_value[sheet.cell(
                        row, col).coordinate] = top_left_value
        return merged_cells_value
